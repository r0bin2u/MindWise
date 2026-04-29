from openpyxl import load_workbook

from mcp_server.tools.excel_writer import HEADERS, append_row


def test_creates_file_with_header_on_first_write(tmp_path):
    target = tmp_path / "log.xlsx"
    assert not target.exists()

    r = append_row("U1", "我很难过", "低落", 1.7, "需关注", path=target)

    assert r["ok"]
    assert target.exists()
    wb = load_workbook(target)
    ws = wb.active
    # header row 1, data row 2
    assert [c.value for c in ws[1]] == HEADERS
    assert ws[2][0].value == "U1"
    assert ws[2][1].value == "我很难过"
    assert ws[2][2].value == "低落"
    assert ws[2][3].value == 1.7
    assert ws[2][4].value == "需关注"
    assert ws[2][5].value is not None  # timestamp was stamped


def test_appends_without_duplicating_header(tmp_path):
    target = tmp_path / "log.xlsx"
    append_row("U1", "a", "正常", 0.0, "正常", path=target)
    append_row("U2", "b", "焦虑", 2.1, "高风险", path=target)
    append_row("U3", "c", "低落", 1.5, "需关注", path=target)

    wb = load_workbook(target)
    ws = wb.active
    assert ws.max_row == 4  # 1 header + 3 data
    # row 1 is the header, rows 2-4 are the three appends in order
    assert ws[2][0].value == "U1"
    assert ws[3][0].value == "U2"
    assert ws[4][0].value == "U3"


def test_honors_explicit_timestamp(tmp_path):
    target = tmp_path / "log.xlsx"
    append_row("U1", "x", "焦虑", 1.8, "需关注", timestamp="2026-04-20 21:15:22", path=target)
    wb = load_workbook(target)
    assert wb.active[2][5].value == "2026-04-20 21:15:22"


def test_creates_parent_dir(tmp_path):
    target = tmp_path / "nested" / "sub" / "log.xlsx"
    append_row("U1", "x", "正常", 0.0, "正常", path=target)
    assert target.exists()
