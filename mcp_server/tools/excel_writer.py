"""Excel log tool — append one consultation record per invocation.

Sheet columns:
    用户ID | 对话内容 | 情绪标签 | 情绪总分 | 风险等级 | 对话时间
    (user_id | message | emotion_label | score | risk_level | timestamp)

The file is created with a header row on first use. Subsequent writes
just append. openpyxl is the standard lib here; we reload-and-rewrite
the whole workbook each call because openpyxl has no true append mode —
for the QPS this system faces (per turn, not per message) that's fine.

Kept as a pure function so it can be called either through the MCP
server (production) or directly from tests and the in-process client
(latency-sensitive on_turn_end path).
"""
from __future__ import annotations

import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.config import settings


HEADERS = ["用户ID", "对话内容", "情绪标签", "情绪总分", "风险等级", "对话时间"]

# Single-process lock so concurrent on_turn_end fires don't clobber each
# other. Multi-process deployments should switch to a file lock or a db.
_write_lock = threading.Lock()


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "consultations"
    ws.append(HEADERS)
    return wb


def append_row(
    user_id: str,
    content: str,
    emotion_label: str,
    score: float,
    risk_level: str,
    timestamp: str | None = None,
    path: str | Path | None = None,
) -> dict[str, Any]:
    """Append a single consultation record. Idempotent on file creation.

    Returns {"ok": True, "row": <1-indexed row number>}.
    Raises on IOError; caller (orchestrator) swallows via BackgroundTasks
    so a failed write never takes down the chat response.
    """
    target = Path(path or settings.excel_log_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    ts = timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [user_id, content, emotion_label, float(score), risk_level, ts]

    with _write_lock:
        wb = _ensure_workbook(target)
        ws = wb.active
        ws.append(row)
        written_row = ws.max_row
        wb.save(target)

    return {"ok": True, "row": written_row, "path": str(target)}
